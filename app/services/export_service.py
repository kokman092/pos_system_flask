import csv
import io
from datetime import datetime, timedelta
from flask import Response
from sqlalchemy import func, cast, Date
from extensions import db
from app.models import Branch, Order, OrderItem, Payment, MenuItem, Category, Ingredient, WasteLog

# ReportLab Imports
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

# Openpyxl Imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================================
# Reusable Helpers
# =========================================================================

def _get_target_date(date_str):
    if date_str:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    return datetime.utcnow().date()


def _get_period_date_range(target_date, period='daily'):
    if period == 'weekly':
        start_day = target_date - timedelta(days=target_date.weekday())
        start = datetime(start_day.year, start_day.month, start_day.day)
        end = start + timedelta(days=7)
    elif period == 'monthly':
        start = datetime(target_date.year, target_date.month, 1)
        if target_date.month == 12:
            end = datetime(target_date.year + 1, 1, 1)
        else:
            end = datetime(target_date.year, target_date.month + 1, 1)
    elif period == 'yearly':
        start = datetime(target_date.year, 1, 1)
        end = datetime(target_date.year + 1, 1, 1)
    else: # daily
        start = datetime(target_date.year, target_date.month, target_date.day)
        end = start + timedelta(days=1)
    return start, end


def stream_csv(headers, data_generator):
    """
    Streams CSV data generator row-by-row with UTF-8 BOM.
    """
    def generate():
        yield '\ufeff'  # BOM for Excel
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        yield output.getvalue()
        
        for row in data_generator:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(row)
            yield output.getvalue()
            
    return generate()


def create_xlsx_response(headers, rows, sheet_name="Report", money_cols=None):
    """
    Generates structured, styled XLSX binary data using openpyxl.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.views.sheetView[0].showGridLines = True
    
    # Elegant Emerald Green Theme
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        
    for row_idx, row in enumerate(rows, 2):
        ws.append(row)
        for col_num, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            
            if money_cols and col_num in money_cols:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
                
    ws.freeze_panes = 'A2'
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =========================================================================
# Query Implementations
# =========================================================================

def _get_revenue_query(branch_id, start_dt, end_dt):
    start_dt_full = datetime(start_dt.year, start_dt.month, start_dt.day)
    end_dt_full = datetime(end_dt.year, end_dt.month, end_dt.day) + timedelta(days=1)
    
    if db.engine.name == 'sqlite':
        date_expr = func.date(Order.opened_at)
    else:
        date_expr = cast(Order.opened_at, Date)
        
    return db.session.query(
        date_expr.label('day'),
        func.count(Order.order_id.distinct()).label('order_count'),
        func.coalesce(func.sum(Payment.amount_cents), 0).label('revenue_cents')
    ).join(Payment, Order.order_id == Payment.order_id)\
     .filter(
         Order.branch_id == branch_id,
         Order.status == 'paid',
         Order.opened_at >= start_dt_full,
         Order.opened_at < end_dt_full,
         Payment.is_voided == 0
     ).group_by(date_expr)\
     .order_by('day')


def _get_payments_query(branch_id, start_dt, end_dt):
    start_dt_full = datetime(start_dt.year, start_dt.month, start_dt.day)
    end_dt_full = datetime(end_dt.year, end_dt.month, end_dt.day) + timedelta(days=1)
    
    return db.session.query(
        Payment.method.label('payment_method'),
        func.count(Payment.payment_id).label('transaction_count'),
        func.coalesce(func.sum(Payment.amount_cents), 0).label('total_cents')
    ).join(Order, Payment.order_id == Order.order_id)\
     .filter(
         Order.branch_id == branch_id,
         Order.status == 'paid',
         Order.opened_at >= start_dt_full,
         Order.opened_at < end_dt_full,
         Payment.is_voided == 0
     ).group_by(Payment.method)\
     .order_by(Payment.method)


def _get_items_query(branch_id, start_dt, end_dt):
    start_dt_full = datetime(start_dt.year, start_dt.month, start_dt.day)
    end_dt_full = datetime(end_dt.year, end_dt.month, end_dt.day) + timedelta(days=1)
    
    return db.session.query(
        MenuItem.name.label('item_name'),
        Category.name.label('category_name'),
        func.sum(OrderItem.quantity).label('qty_sold'),
        func.coalesce(func.sum(OrderItem.quantity * OrderItem.unit_price_cents), 0).label('revenue_cents')
    ).join(OrderItem, MenuItem.item_id == OrderItem.item_id)\
     .join(Order, Order.order_id == OrderItem.order_id)\
     .join(Category, Category.category_id == MenuItem.category_id)\
     .filter(
         Order.branch_id == branch_id,
         Order.status == 'paid',
         Order.opened_at >= start_dt_full,
         Order.opened_at < end_dt_full,
         OrderItem.status != 'cancelled'
     ).group_by(MenuItem.item_id, MenuItem.name, Category.name)\
     .order_by(func.sum(OrderItem.quantity).desc())


def _get_waste_query(branch_id, start_dt, end_dt):
    start_dt_full = datetime(start_dt.year, start_dt.month, start_dt.day)
    end_dt_full = datetime(end_dt.year, end_dt.month, end_dt.day) + timedelta(days=1)
    
    if db.engine.name == 'sqlite':
        date_expr = func.date(WasteLog.created_at)
    else:
        date_expr = cast(WasteLog.created_at, Date)
        
    return db.session.query(
        date_expr.label('day'),
        Ingredient.name.label('ingredient_name'),
        WasteLog.qty.label('qty'),
        WasteLog.reason.label('reason'),
        WasteLog.unit_cost_cents.label('unit_cost_cents'),
        (WasteLog.qty * WasteLog.unit_cost_cents).label('total_cost_cents')
    ).join(Ingredient, WasteLog.ingredient_id == Ingredient.ingredient_id)\
     .filter(
         WasteLog.branch_id == branch_id,
         WasteLog.created_at >= start_dt_full,
         WasteLog.created_at < end_dt_full
     ).order_by(date_expr.desc())


def _get_inventory_query(branch_id):
    return db.session.query(
        Ingredient.name.label('ingredient_name'),
        Ingredient.unit.label('unit'),
        Ingredient.qty_in_stock.label('qty_in_stock'),
        Ingredient.cost_per_unit_cents.label('cost_per_unit_cents'),
        (Ingredient.qty_in_stock * Ingredient.cost_per_unit_cents).label('stock_value_cents')
    ).filter(
        Ingredient.branch_id == branch_id
    ).order_by(Ingredient.name)


# =========================================================================
# Export Actions
# =========================================================================

# --- Revenue Exports ---
def export_revenue_csv(branch_id, start_date, end_date):
    start_dt = _get_target_date(start_date)
    end_dt = _get_target_date(end_date)
    headers = ["date", "order_count", "revenue_cents", "revenue_display"]
    
    query = _get_revenue_query(branch_id, start_dt, end_dt)
    
    def generator():
        for r in query.yield_per(100):
            yield [
                str(r.day),
                r.order_count,
                int(r.revenue_cents),
                f"${r.revenue_cents / 100:.2f}"
            ]
            
    return stream_csv(headers, generator())


def export_revenue_xlsx(branch_id, start_date, end_date):
    start_dt = _get_target_date(start_date)
    end_dt = _get_target_date(end_date)
    headers = ["date", "order_count", "revenue_cents", "revenue_display"]
    
    query = _get_revenue_query(branch_id, start_dt, end_dt).all()
    rows = []
    for r in query:
        rows.append([
            str(r.day),
            r.order_count,
            int(r.revenue_cents) / 100.0,
            f"${r.revenue_cents / 100:.2f}"
        ])
    return create_xlsx_response(headers, rows, sheet_name="Revenue", money_cols=[3])


# --- Payment Exports ---
def export_payments_csv(branch_id, start_date, end_date):
    start_dt = _get_target_date(start_date)
    end_dt = _get_target_date(end_date)
    headers = ["payment_method", "transaction_count", "total_cents", "total_display"]
    
    query = _get_payments_query(branch_id, start_dt, end_dt)
    
    def generator():
        for r in query.yield_per(100):
            yield [
                r.payment_method,
                r.transaction_count,
                int(r.total_cents),
                f"${r.total_cents / 100:.2f}"
            ]
            
    return stream_csv(headers, generator())


def export_payments_xlsx(branch_id, start_date, end_date):
    start_dt = _get_target_date(start_date)
    end_dt = _get_target_date(end_date)
    headers = ["payment_method", "transaction_count", "total_cents", "total_display"]
    
    query = _get_payments_query(branch_id, start_dt, end_dt).all()
    rows = []
    for r in query:
        rows.append([
            r.payment_method,
            r.transaction_count,
            int(r.total_cents) / 100.0,
            f"${r.total_cents / 100:.2f}"
        ])
    return create_xlsx_response(headers, rows, sheet_name="Payments", money_cols=[3])


# --- Item Sales Exports ---
def export_items_csv(branch_id, start_date, end_date):
    start_dt = _get_target_date(start_date)
    end_dt = _get_target_date(end_date)
    headers = ["item_name", "category_name", "qty_sold", "revenue_cents", "revenue_display"]
    
    query = _get_items_query(branch_id, start_dt, end_dt)
    
    def generator():
        for r in query.yield_per(100):
            yield [
                r.item_name,
                r.category_name,
                r.qty_sold,
                int(r.revenue_cents),
                f"${r.revenue_cents / 100:.2f}"
            ]
            
    return stream_csv(headers, generator())


def export_items_xlsx(branch_id, start_date, end_date):
    start_dt = _get_target_date(start_date)
    end_dt = _get_target_date(end_date)
    headers = ["item_name", "category_name", "qty_sold", "revenue_cents", "revenue_display"]
    
    query = _get_items_query(branch_id, start_dt, end_dt).all()
    rows = []
    for r in query:
        rows.append([
            r.item_name,
            r.category_name,
            r.qty_sold,
            int(r.revenue_cents) / 100.0,
            f"${r.revenue_cents / 100:.2f}"
        ])
    return create_xlsx_response(headers, rows, sheet_name="Item Sales", money_cols=[4])


# --- Waste Exports ---
def export_waste_csv(branch_id, start_date, end_date):
    start_dt = _get_target_date(start_date)
    end_dt = _get_target_date(end_date)
    headers = ["date", "ingredient_name", "qty", "reason", "unit_cost_cents", "total_cost_cents", "total_cost_display"]
    
    query = _get_waste_query(branch_id, start_dt, end_dt)
    
    def generator():
        for r in query.yield_per(100):
            yield [
                str(r.day),
                r.ingredient_name,
                float(r.qty),
                r.reason,
                int(r.unit_cost_cents),
                int(r.total_cost_cents),
                f"${r.total_cost_cents / 100:.2f}"
            ]
            
    return stream_csv(headers, generator())


def export_waste_xlsx(branch_id, start_date, end_date):
    start_dt = _get_target_date(start_date)
    end_dt = _get_target_date(end_date)
    headers = ["date", "ingredient_name", "qty", "reason", "unit_cost_cents", "total_cost_cents", "total_cost_display"]
    
    query = _get_waste_query(branch_id, start_dt, end_dt).all()
    rows = []
    for r in query:
        rows.append([
            str(r.day),
            r.ingredient_name,
            float(r.qty),
            r.reason,
            int(r.unit_cost_cents) / 100.0,
            int(r.total_cost_cents) / 100.0,
            f"${r.total_cost_cents / 100:.2f}"
        ])
    return create_xlsx_response(headers, rows, sheet_name="Waste", money_cols=[5, 6])


# --- Inventory Exports ---
def export_inventory_csv(branch_id):
    headers = ["ingredient_name", "unit", "qty_in_stock", "cost_per_unit_cents", "stock_value_cents", "stock_value_display"]
    
    query = _get_inventory_query(branch_id)
    
    def generator():
        for r in query.yield_per(100):
            yield [
                r.ingredient_name,
                r.unit,
                float(r.qty_in_stock),
                int(r.cost_per_unit_cents),
                int(r.stock_value_cents),
                f"${r.stock_value_cents / 100:.2f}"
            ]
            
    return stream_csv(headers, generator())


def export_inventory_xlsx(branch_id):
    headers = ["ingredient_name", "unit", "qty_in_stock", "cost_per_unit_cents", "stock_value_cents", "stock_value_display"]
    
    query = _get_inventory_query(branch_id).all()
    rows = []
    for r in query:
        rows.append([
            r.ingredient_name,
            r.unit,
            float(r.qty_in_stock),
            int(r.cost_per_unit_cents) / 100.0,
            int(r.stock_value_cents) / 100.0,
            f"${r.stock_value_cents / 100:.2f}"
        ])
    return create_xlsx_response(headers, rows, sheet_name="Inventory", money_cols=[4, 5])


# =========================================================================
# PDF Exporter (ReportLab)
# =========================================================================

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#64748B"))
        
        # Timestamp
        timestamp_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        self.drawString(54, 36, f"Exported on: {timestamp_str}")
        
        # Page Number
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(A4[0] - 54, 36, page_text)
        
        # Footer border line
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(54, 50, A4[0] - 54, 50)
        self.restoreState()


def _build_pdf_document(branch_name, title, date_range, kpis, payment_methods, top_items):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=72
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0F172A")
    )
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#64748B")
    )
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#10B981"),
        spaceBefore=14,
        spaceAfter=8
    )
    th_style = ParagraphStyle(
        'TH',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#FFFFFF")
    )
    td_style = ParagraphStyle(
        'TD',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#0F172A")
    )
    
    kpi_title_style = ParagraphStyle(
        'KPITitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#64748B"),
        alignment=1
    )
    kpi_value_style = ParagraphStyle(
        'KPIValue',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=16,
        textColor=colors.HexColor("#10B981"),
        alignment=1
    )
    
    story = []
    
    # Header block
    story.append(Paragraph(f"AURA POS - {title.upper()}", title_style))
    story.append(Paragraph(f"Branch: {branch_name}  |  Period: {date_range}", subtitle_style))
    story.append(Spacer(1, 15))
    
    # KPI Grid (4 Columns table)
    kpi_data = [
        [
            Paragraph("TOTAL REVENUE", kpi_title_style),
            Paragraph("TOTAL ORDERS", kpi_title_style),
            Paragraph("AVG TICKET SIZE", kpi_title_style),
            Paragraph("TOTAL COVERS", kpi_title_style)
        ],
        [
            Paragraph(kpis.get('revenue', '$0.00'), kpi_value_style),
            Paragraph(str(kpis.get('orders', 0)), kpi_value_style),
            Paragraph(kpis.get('avg_order', '$0.00'), kpi_value_style),
            Paragraph(str(kpis.get('covers', 0)), kpi_value_style)
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[120, 120, 120, 120])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # Table headers style helper
    th_bg = colors.HexColor("#10B981")
    border_color = colors.HexColor("#E2E8F0")
    
    # 1. Payment Breakdown Table
    story.append(Paragraph("Payment Method Breakdown", section_title_style))
    pm_rows = [[Paragraph("Payment Method", th_style), Paragraph("Transactions", th_style), Paragraph("Total Amount", th_style)]]
    for pm in payment_methods:
        pm_rows.append([
            Paragraph(pm['method'].replace('_', ' ').capitalize(), td_style),
            Paragraph(str(pm['count']), td_style),
            Paragraph(f"${pm['total_cents'] / 100:.2f}", td_style)
        ])
    
    pm_table = Table(pm_rows, colWidths=[200, 140, 140])
    pm_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), th_bg),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, border_color),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F8FAFC")])
    ]))
    story.append(pm_table)
    story.append(Spacer(1, 15))
    
    # 2. Top Selling Items Table
    story.append(Paragraph("Top Selling Items", section_title_style))
    item_rows = [[Paragraph("Item Name", th_style), Paragraph("Qty Sold", th_style), Paragraph("Revenue Generated", th_style)]]
    for item in top_items:
        item_rows.append([
            Paragraph(item['name'], td_style),
            Paragraph(str(item['qty_sold']), td_style),
            Paragraph(f"${item['revenue_cents'] / 100:.2f}", td_style)
        ])
        
    item_table = Table(item_rows, colWidths=[240, 120, 120])
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), th_bg),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, border_color),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F8FAFC")])
    ]))
    story.append(item_table)
    
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.getvalue()


def export_daily_pdf(branch_id, date_str):
    branch = db.session.get(Branch, branch_id)
    if not branch:
        raise ValueError("Branch not found")
    target_date = _get_target_date(date_str)
    
    # Fetch aggregates from report_service logic
    import app.services.report_service as report_service
    data = report_service.daily_summary(branch_id, date_str, period='daily')
    
    kpis = {
        'revenue': f"${data['total_revenue_cents'] / 100:.2f}",
        'orders': data['total_orders'],
        'avg_order': f"${data['avg_order_cents'] / 100:.2f}",
        'covers': data['total_customers']
    }
    
    return _build_pdf_document(
        branch_name=branch.name,
        title="Daily Summary Report",
        date_range=target_date.strftime("%B %d, %Y"),
        kpis=kpis,
        payment_methods=data['payment_methods'],
        top_items=data['top_items']
    )


def export_shift_pdf(branch_id, shift, date_str):
    branch = db.session.get(Branch, branch_id)
    if not branch:
        raise ValueError("Branch not found")
    target_date = _get_target_date(date_str)
    
    import app.services.report_service as report_service
    shift_data = report_service.shift_report(branch_id, shift, date_str, period='daily')
    daily_data = report_service.daily_summary(branch_id, date_str, period='daily')
    
    # Filter payment methods and items by shift boundaries manually to reflect correct shift context
    start_dt, end_dt = _get_period_date_range(target_date, 'daily')
    if shift == 'morning':
        end_dt = start_dt + timedelta(hours=12)
    elif shift == 'afternoon':
        start_dt = start_dt + timedelta(hours=12)
        end_dt = start_dt + timedelta(hours=6)
    elif shift == 'night':
        start_dt = start_dt + timedelta(hours=18)

    # Shift specific payment methods
    pm_query = db.session.query(
        Payment.method,
        func.count(Payment.payment_id).label('count'),
        func.coalesce(func.sum(Payment.amount_cents), 0).label('total')
    ).join(Order, Payment.order_id == Order.order_id)\
     .filter(
         Order.branch_id == branch_id,
         Order.status == 'paid',
         Order.opened_at >= start_dt,
         Order.opened_at < end_dt,
         Payment.is_voided == 0
     ).group_by(Payment.method).all()
    
    payment_methods = [{"method": m, "count": cnt, "total_cents": int(tot)} for m, cnt, tot in pm_query]
    
    # Shift specific top items
    top_query = db.session.query(
        MenuItem.name,
        func.sum(OrderItem.quantity).label('qty_sold'),
        func.coalesce(func.sum(OrderItem.quantity * OrderItem.unit_price_cents), 0).label('revenue')
    ).join(OrderItem, MenuItem.item_id == OrderItem.item_id)\
     .join(Order, Order.order_id == OrderItem.order_id)\
     .filter(
         Order.branch_id == branch_id,
         Order.status == 'paid',
         Order.opened_at >= start_dt,
         Order.opened_at < end_dt,
         OrderItem.status != 'cancelled'
     ).group_by(MenuItem.item_id, MenuItem.name)\
     .order_by(func.sum(OrderItem.quantity).desc())\
     .limit(5).all()
     
    top_items = [{"name": n, "qty_sold": q, "revenue_cents": int(rev)} for n, q, rev in top_query]
    
    total_orders = shift_data['total_orders']
    revenue_cents = shift_data['total_revenue_cents']
    avg_order_cents = revenue_cents // total_orders if total_orders > 0 else 0
    
    # Covers for shift
    covers = db.session.query(
        func.coalesce(func.sum(Order.pax), 0)
    ).filter(
        Order.branch_id == branch_id,
        Order.status == 'paid',
        Order.opened_at >= start_dt,
        Order.opened_at < end_dt
    ).scalar() or 0
    
    kpis = {
        'revenue': f"${revenue_cents / 100:.2f}",
        'orders': total_orders,
        'avg_order': f"${avg_order_cents / 100:.2f}",
        'covers': int(covers)
    }
    
    return _build_pdf_document(
        branch_name=branch.name,
        title=f"Shift Performance Report ({shift.capitalize()})",
        date_range=target_date.strftime("%B %d, %Y"),
        kpis=kpis,
        payment_methods=payment_methods,
        top_items=top_items
    )
